from tkinter import *
from openpyxl import load_workbook
from dadosXLSX import Dados



def cadastroFuncionarios():

    # Funções
    def cadastro():
        book = load_workbook('Cadastro de técnicos.xlsx')
        print(book.sheetnames)

        # Selecionar a página do exel
        nome_do_tecnico_pagina = book['Nome']
        cpf_pagina = book['CPF']
        telefone_pagina = book['Telefone']
        turno_pagina = book['Turno']
        nome_da_equipe_pagina = book['NomeEquipe']

        # Adicionar elementos na página



        nome_do_tecnico_pagina.append([nome_do_tecnico.get()])
        cpf_pagina.append([cpf.get()])
        telefone_pagina.append([telefone.get()])
        turno_pagina.append([turno.get()])
        nome_da_equipe_pagina.append([nome_da_equipe.get()])
        book.save('Cadastro de técnicos.xlsx')

        master.destroy()
    

    def retorno():

        master.destroy()

    def click_mouse(retorno):
           print(retorno)
           print(f'x : {retorno.x} | y : {retorno.y} | geo : {master.geometry()}')

    #AKI SEU CODIGO

    master =  Toplevel()
    #master = Tk()
    master.title("***Cadastro do colaborador****")
    master.iconbitmap(default="")#trocar o ícone
    master.geometry("1200x870+352+79") # Largura x Altura + dist esquerda + dist topo
    #900x600+591+215  
    master.resizable(width=1, height=1)


    #Importar imagens
    img_principal = PhotoImage(file="img_cadastro_colaboradores.png")
    img_cadastrar = PhotoImage(file="")#botao_cadastrar.png
    img_retornar = PhotoImage(file="") #botao_retornar.png

    # funções
    # Função para saber as coordenandas dos botões e o tamanho da tela principal
    #def clique_esq_mouse(retorno):
    #    print(f'x: {retorno.x} | y: {retorno.y} Geo: {master.geometry()}')

    # Criação de Labels
    lab_principal = Label(master, image=img_principal)
    lab_principal.place(x=0, y=0)
    lab_principal.pack()

    # Criação de caixas de entrada
    nome_do_tecnico = Entry(master , bd=2 , font=("Calibri",15) , justify=CENTER)
    nome_do_tecnico.place(width=584 , height=34, x=438 , y=206)

    cpf = Entry(master , bd=2 , font=("Calibri",15) , justify=CENTER)
    cpf.place(width=584 , height=36 , x=438 , y=300)

    telefone = Entry(master , bd=2 , font=("Calibri",15) , justify=CENTER)
    telefone.place(width=584 , height=36 , x=438 , y=390)

    turno = Entry(master , bd=2 , font=("Calibri",15) , justify=CENTER)
    turno.place(width=584 , height=36 , x=438 , y=481)

    nome_da_equipe = Entry(master , bd=2 , font=("Calibri",15) , justify=CENTER)
    nome_da_equipe.place(width=584 , height=36 , x=438 , y=565)

    # Criação de Botões
    cadastrar = Button(master ,text="confimar", width=16, height=2, bg="orange" , command = cadastro )
    cadastrar.place(x=700 , y=814)
    #place(width=340, height=100 , x=50 , y=695)
    retornar = Button(master ,text="retornar",  width=16, height=2, bg="orange" , command = retorno)
    retornar.place(x=948 , y=814)
    #place(width=166, height=49 , x=450 , y=769)
    #eventos
    master.bind("<Button-1>",click_mouse)
    # Salvar a planilha

    #master.mainloop()

#cadastroFuncionarios()